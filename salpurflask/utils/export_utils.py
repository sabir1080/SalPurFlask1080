"""Export utilities for CSV and Excel generation."""

import csv
from io import StringIO, BytesIO

from flask import current_app, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from salpurflask.utils.helpers import now_local


def write_csv_header(writer, report_title, start_date_str=None, end_date_str=None, extra_info=None):
    """Write standard CSV report header rows.

    Args:
        writer: csv.writer object
        report_title: Title of the report
        start_date_str: Optional start date string
        end_date_str: Optional end date string
        extra_info: Optional extra information line
    """
    company = current_app.config.get("COMPANY_NAME", "Company")
    tagline = current_app.config.get("COMPANY_TAGLINE", "")
    writer.writerow([company, tagline])
    writer.writerow(["Report:", report_title])
    if start_date_str and end_date_str:
        writer.writerow(["Period:", f"{start_date_str}  to  {end_date_str}"])
    if extra_info:
        writer.writerow(["Info:", extra_info])
    writer.writerow(["Generated On:", now_local().strftime("%Y-%m-%d %H:%M")])
    writer.writerow([])


def csv_response(filename, title, col_headers, rows, start_date_str=None, end_date_str=None, extra_info=None):
    """Build a CSV entirely in memory and return it as a download.

    No shared file on disk, so concurrent exports from different users/tabs
    can never race or overwrite each other.

    Args:
        filename: Download filename
        title: Report title
        col_headers: List of column header strings
        rows: List of row data (list of lists)
        start_date_str: Optional start date
        end_date_str: Optional end date
        extra_info: Optional extra information

    Returns:
        Flask Response with CSV data
    """
    buf = StringIO()
    writer = csv.writer(buf)
    write_csv_header(writer, title, start_date_str, end_date_str, extra_info)
    writer.writerow(col_headers)
    writer.writerows(rows)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def excel_response(filename, title, col_headers, rows, start_date_str=None, end_date_str=None, extra_info=None):
    """Create a styled .xlsx file and return as a Flask file download response.

    Args:
        filename: Download filename
        title: Report title
        col_headers: List of column header strings
        rows: List of row data (list of lists)
        start_date_str: Optional start date
        end_date_str: Optional end date
        extra_info: Optional extra information

    Returns:
        Flask file download response with Excel data
    """
    company = current_app.config.get("COMPANY_NAME", "Company")
    tagline = current_app.config.get("COMPANY_TAGLINE", "")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Metadata rows
    r = 1
    ws.cell(row=r, column=1, value=company)
    ws.cell(row=r, column=2, value=tagline)
    ws["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    r += 1

    ws.cell(row=r, column=1, value="Report:")
    ws.cell(row=r, column=2, value=title)
    r += 1

    if start_date_str and end_date_str:
        ws.cell(row=r, column=1, value="Period:")
        ws.cell(row=r, column=2, value=f"{start_date_str}  to  {end_date_str}")
        r += 1

    if extra_info:
        ws.cell(row=r, column=1, value="Info:")
        ws.cell(row=r, column=2, value=extra_info)
        r += 1

    ws.cell(row=r, column=1, value="Generated On:")
    ws.cell(row=r, column=2, value=now_local().strftime("%Y-%m-%d %H:%M"))
    r += 1

    r += 1  # Blank row

    # Column header row
    header_row_num = r
    for col_i, h in enumerate(col_headers, 1):
        cell = ws.cell(row=header_row_num, column=col_i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    # Data rows
    for row_data in rows:
        for col_i, val in enumerate(row_data, 1):
            ws.cell(row=r, column=col_i, value=val)
        r += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


__all__ = ['csv_response', 'excel_response', 'write_csv_header']
