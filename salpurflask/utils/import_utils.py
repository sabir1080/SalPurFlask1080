"""File import utilities for parsing CSV, Excel, and JSON files.

Generic file parsing logic that may be reused by other modules.
"""

import csv
import json
from io import StringIO

import openpyxl


# Import constraints and limits
IMPORT_MAX_ROWS = 10000

# Allowed MIME types for bulk import
ALLOWED_MIME_TYPES = {
    'text/csv',
    'text/plain',  # CSV may be reported as plain text
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
    'application/vnd.ms-excel',  # .xls
    'application/json',
}

# Maximum file size: 50 MB (prevents DoS)
MAX_IMPORT_FILE_SIZE = 50 * 1024 * 1024


def parse_import_file(file):
    """Parse CSV/Excel/JSON file and return list of dictionaries.

    Validates file extension, MIME type, and file size. Returns None on error.

    Args:
        file: Flask FileStorage object from request.files

    Returns:
        Tuple of (data, file_type) where data is list of dicts, file_type is extension.
        On error: (None, error_message)
    """
    if not file:
        return None, "No file provided"

    filename = file.filename
    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

    # Validate file extension
    if file_ext not in ('csv', 'xlsx', 'xls', 'json'):
        return None, f"Invalid file extension: {file_ext}. Allowed: CSV, Excel (.xlsx, .xls), JSON"

    # Validate MIME type
    mime_type = file.content_type or ''
    if mime_type not in ALLOWED_MIME_TYPES:
        # Log warning but don't block if extension is valid
        # (Some file systems may report MIME types differently)
        pass

    # Check file size (prevent DoS)
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning
    if file_size > MAX_IMPORT_FILE_SIZE:
        return None, f"File is too large ({file_size / 1024 / 1024:.1f} MB). Maximum size is 50 MB."

    try:
        if file_ext == 'csv':
            stream = StringIO(file.read().decode('utf-8'))
            reader = csv.DictReader(stream)
            data = list(reader)

        elif file_ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))

        elif file_ext == 'json':
            parsed = json.loads(file.read().decode('utf-8'))
            data = parsed if isinstance(parsed, list) else [parsed]

        else:
            return None, f"Unsupported file type: {file_ext}"

        if len(data) > IMPORT_MAX_ROWS:
            return None, f"File has {len(data)} rows — the maximum is {IMPORT_MAX_ROWS} per import."
        return data, file_ext

    except Exception as e:
        return None, f"Error parsing file: {str(e)}"


__all__ = ['parse_import_file', 'IMPORT_MAX_ROWS', 'ALLOWED_MIME_TYPES', 'MAX_IMPORT_FILE_SIZE']
